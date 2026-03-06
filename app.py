import streamlit as st
import openpyxl
import io
import re

def load_content_master_lookup(file_buffer):
    wb = openpyxl.load_workbook(file_buffer, data_only=True)
    ws = wb.active
    headers = {cell.value.strip(): col_idx for col_idx, cell in enumerate(ws[1], start=1) if cell.value}
    lookup = {}
    required_fields = {"Title": "Final Product Title", "HTML content": "HTML Content", "Brand Name": "Brand Name", "Colour": "Final Color", "Care Instruction": "Care Instruction"}
    for row in range(2, ws.max_row + 1):
        bz_code = ws.cell(row=row, column=headers.get("BZ CODE", 0)).value
        if bz_code:
            bz_code = str(bz_code).strip()
            lookup[bz_code] = {k: ws.cell(row=row, column=headers.get(v, 0)).value for k, v in required_fields.items()}
    return lookup

def load_master_lookup(file_buffer):
    wb = openpyxl.load_workbook(file_buffer, data_only=True)
    ws = wb.active
    headers = {cell.value.strip(): col_idx for col_idx, cell in enumerate(ws[1], start=1) if cell.value}
    lookup = {}
    for row in range(2, ws.max_row + 1):
        article = ws.cell(row=row, column=headers.get("Article", 0)).value
        if article:
            article = str(article).strip()
            lookup[article] = {"Size": ws.cell(row=row, column=headers.get("Size", 0)).value, "New MRP": ws.cell(row=row, column=headers.get("NEW MRP", headers.get("New MRP", 0))).value, "Old MRP": ws.cell(row=row, column=headers.get("OLD MRP", headers.get("Old MRP", 0))).value, "EAN/UPC": ws.cell(row=row, column=headers.get("EAN/UPC", 0)).value, "Country": ws.cell(row=row, column=headers.get("Country", 0)).value, "Dimension": ws.cell(row=row, column=headers.get("Dimension", 0)).value, "Article": article}
    return lookup

def process_excel_data(products_buffer, content_buffer, master_buffer):
    content_lookup = load_content_master_lookup(content_buffer)
    master_lookup = load_master_lookup(master_buffer)
    wb = openpyxl.load_workbook(products_buffer)
    ws = wb.active
    headers = {cell.value.strip(): col_idx for col_idx, cell in enumerate(ws[1], start=1) if cell.value}
    sku_col = headers.get("SKU")
    mapping = [("Title", "content", "Title"), ("Body (HTML)", "content", "HTML content"), ("Vendor", "content", "Brand Name"), ("Option1 Value", "content", "Colour"), ("Option2 Value", "master", "Size"), ("Variant SKU", "master", "Article"), ("Variant Price", "master", "New MRP"), ("Variant Compare At Price", "master", "Old MRP"), ("Variant Barcode", "master", "EAN/UPC"), ("Size (product.metafields.custom.size)", "master", "Size"), ("Care Instruction (product.metafields.my_fields.care_instruction)", "content", "Care Instruction"), ("Country of origin (product.metafields.my_fields.country_of_origin)", "master", "Country"), ("Dimensions (product.metafields.my_fields.specifications)", "master", "Dimension")]
    
    for row in range(2, ws.max_row + 1):
        sku = str(ws.cell(row=row, column=sku_col).value or "").strip()
        if not sku: continue
        c_data, m_data = content_lookup.get(sku), master_lookup.get(sku)
        
        country_of_origin = str(m_data.get("Country") or "").strip().title() if m_data else ""
        
        for prod_col, src, src_fld in mapping:
            idx = headers.get(prod_col)
            if not idx: continue
            val = c_data.get(src_fld) if src == "content" and c_data else (m_data.get(src_fld) if src == "master" and m_data else None)
            if val is not None:
                if isinstance(val, str):
                    val = val.strip()
                    if prod_col in ["Vendor", "Option1 Value", "Option2 Value", "Size (product.metafields.custom.size)", "Country of origin (product.metafields.my_fields.country_of_origin)"]: val = val.title()
                    if prod_col in ["Option2 Value", "Size (product.metafields.custom.size)"] and val.lower() == "x-large": val = "Extra Large"
                ws.cell(row=row, column=idx, value=val)
        
        m_idx = headers.get("Manufacturer Details (product.metafields.my_fields.manufacturer_details)")
        if m_idx and country_of_origin:
            if country_of_origin == "India":
                ws.cell(row=row, column=m_idx, value="<p><strong>Manufatured: </strong>Bagzone Lifestyles Private Limited 401, Ackruti Oppo. Ackruti Centre Point, Central Road, MIDC, Andheri East, Mumbai-400093.</p>")
            elif country_of_origin == "China":
                ws.cell(row=row, column=m_idx, value="<p><strong>Imported & Marketed By: </strong>Bagzone Lifestyles Private Limited 401, Ackruti Star, Oppo. Ackruti Centre Point, Central Road, MIDC, Andheri East, Mumbai-400093.</p>")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

import os

# Get the directory where the app is located
APP_DIR = os.path.dirname(os.path.abspath(__file__))
CONTENT_FILE = os.path.join(APP_DIR, "content_master.xlsx")
GS_FILE = os.path.join(APP_DIR, "gs.xlsx")

st.title("Product Data Populator")

# Check if reference files exist
content_exists = os.path.exists(CONTENT_FILE)
gs_exists = os.path.exists(GS_FILE)

if not content_exists or not gs_exists:
    st.error("Reference files missing! Please ensure these files are in the app folder:")
    if not content_exists:
        st.warning("❌ content_master.xlsx - NOT FOUND")
    else:
        st.success("✅ content_master.xlsx - Found")
    if not gs_exists:
        st.warning("❌ gs.xlsx - NOT FOUND")
    else:
        st.success("✅ gs.xlsx - Found")
else:
    st.success("✅ Reference files loaded: content_master.xlsx, gs.xlsx")
    
    prod_file = st.file_uploader("Upload products.xlsx", type=["xlsx"])
    
    if prod_file:
        if st.button("Process Data"):
            result = process_excel_data(prod_file, CONTENT_FILE, GS_FILE)
            st.success("Processing Complete!")
            st.download_button("Download Result", data=result, file_name="new6.xlsx")
