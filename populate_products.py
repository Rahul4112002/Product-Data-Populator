"""
Populate products.xlsx by looking up data from master.xlsx and content-master-1_transformed.xlsx.
Uses SKU as the primary key to match against BZ CODE (content-master) and Article (master).
Outputs: products_filled.xlsx + summary report in console.
"""

import openpyxl
from collections import defaultdict


def load_content_master_lookup(filepath):
    """Load content-master-1_transformed.xlsx and build a lookup dict keyed by BZ CODE."""
    print(f"Loading content-master: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Build header map
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[cell.value.strip()] = col_idx

    required_cols = ["BZ CODE", "Final Product Title", "HTML Content", "Brand Name", "Final Color", "Care Instruction"]
    for col_name in required_cols:
        if col_name not in headers:
            print(f"  WARNING: Column '{col_name}' not found in content-master. Available: {list(headers.keys())}")

    lookup = {}
    for row in range(2, ws.max_row + 1):
        bz_code = ws.cell(row=row, column=headers.get("BZ CODE", 0)).value
        if bz_code is None:
            continue
        bz_code = str(bz_code).strip()
        if bz_code and bz_code not in lookup:
            lookup[bz_code] = {
                "Title": ws.cell(row=row, column=headers.get("Final Product Title", 0)).value,
                "HTML content": ws.cell(row=row, column=headers.get("HTML Content", 0)).value,
                "Brand Name": ws.cell(row=row, column=headers.get("Brand Name", 0)).value,
                "Colour": ws.cell(row=row, column=headers.get("Final Color", 0)).value,
                "Care Instruction": ws.cell(row=row, column=headers.get("Care Instruction", 0)).value,
            }

    wb.close()
    print(f"  Loaded {len(lookup)} unique BZ CODE entries.")
    return lookup


def load_master_lookup(filepath):
    """Load master.xlsx and build a lookup dict keyed by Article."""
    print(f"Loading master: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Build header map
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[cell.value.strip()] = col_idx

    required_cols = ["Article", "Size", "New MRP", "Old MRP", "EAN/UPC", "Country", "Dimension"]
    for col_name in required_cols:
        # Check for close matches (e.g. "Country " with trailing space)
        matched = False
        for h in headers:
            if h.strip().lower() == col_name.strip().lower():
                matched = True
                break
        if not matched:
            print(f"  WARNING: Column '{col_name}' not found in master. Available: {list(headers.keys())}")

    # Normalize header keys (strip whitespace)
    normalized_headers = {k.strip(): v for k, v in headers.items()}

    # Find the first "Article" column (there are duplicates at col 2 and col 58)
    article_col = normalized_headers.get("Article")

    lookup = {}
    for row in range(2, ws.max_row + 1):
        article = ws.cell(row=row, column=article_col).value
        if article is None:
            continue
        article = str(article).strip()
        if article and article not in lookup:
            lookup[article] = {
                "Size": ws.cell(row=row, column=normalized_headers.get("Size", 0)).value,
                "New MRP": ws.cell(row=row, column=normalized_headers.get("NEW MRP", normalized_headers.get("New MRP", 0))).value,
                "Old MRP": ws.cell(row=row, column=normalized_headers.get("OLD MRP", normalized_headers.get("Old MRP", 0))).value,
                "EAN/UPC": ws.cell(row=row, column=normalized_headers.get("EAN/UPC", 0)).value,
                "Country": ws.cell(row=row, column=normalized_headers.get("Country", 0)).value,
                "Dimension": ws.cell(row=row, column=normalized_headers.get("Dimension", 0)).value,
                "Article": article,
            }

    wb.close()
    print(f"  Loaded {len(lookup)} unique Article entries.")
    return lookup


def populate_products(products_path, content_master_path, master_path, output_path):
    """Main function to populate products.xlsx with lookup data."""

    # Load lookup dictionaries
    content_lookup = load_content_master_lookup(content_master_path)
    master_lookup = load_master_lookup(master_path)

    # Load products workbook
    print(f"\nLoading products: {products_path}")
    wb = openpyxl.load_workbook(products_path)
    ws = wb.active

    # Build products header map
    product_headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            product_headers[cell.value.strip()] = col_idx

    # Define the column mapping
    # Format: (products_column_name, source, source_field)
    # source = "content-master" or "master"
    column_mapping = [
        ("Title", "content-master", "Title"),
        ("Body (HTML)", "content-master", "HTML content"),
        ("Vendor", "content-master", "Brand Name"),
        ("Option1 Value", "content-master", "Colour"),
        ("Option2 Value", "master", "Size"),
        ("Variant SKU", "master", "Article"),
        ("Variant Price", "master", "New MRP"),
        ("Variant Compare At Price", "master", "Old MRP"),
        ("Variant Barcode", "master", "EAN/UPC"),
        ("Size (product.metafields.custom.size)", "master", "Size"),
        ("Care Instruction (product.metafields.my_fields.care_instruction)", "content-master", "Care Instruction"),
        ("Country of origin (product.metafields.my_fields.country_of_origin)", "master", "Country"),
        ("Dimensions (product.metafields.my_fields.specifications)", "master", "Dimension"),
    ]

    # Verify all target columns exist in products sheet
    for prod_col, source, source_field in column_mapping:
        if prod_col not in product_headers:
            print(f"  WARNING: Products column '{prod_col}' not found!")

    # Get SKU column index
    sku_col = product_headers.get("SKU")
    if not sku_col:
        print("ERROR: SKU column not found in products sheet!")
        return

    # Stats tracking
    total_rows = 0
    rows_with_sku = 0
    content_matches = 0
    master_matches = 0
    no_match_rows = 0

    # Process each row
    print(f"\nProcessing {ws.max_row - 1} data rows...")
    for row in range(2, ws.max_row + 1):
        total_rows += 1
        sku_value = ws.cell(row=row, column=sku_col).value

        if sku_value is None or str(sku_value).strip() == "":
            continue

        sku = str(sku_value).strip()
        rows_with_sku += 1

        # Lookup from content-master
        content_data = content_lookup.get(sku)
        has_content_match = content_data is not None
        if has_content_match:
            content_matches += 1

        # Lookup from master
        master_data = master_lookup.get(sku)
        has_master_match = master_data is not None
        if has_master_match:
            master_matches += 1

        if not has_content_match and not has_master_match:
            no_match_rows += 1

        # Fill in the columns
        for prod_col, source, source_field in column_mapping:
            target_col_idx = product_headers.get(prod_col)
            if not target_col_idx:
                continue

            value = None
            if source == "content-master" and content_data:
                value = content_data.get(source_field)
            elif source == "master" and master_data:
                value = master_data.get(source_field)

            # Only write if we have a value (preserve existing data otherwise)
            if value is not None:
                ws.cell(row=row, column=target_col_idx, value=value)

    # Save the output file
    wb.save(output_path)
    wb.close()
    print(f"\nSaved filled products to: {output_path}")

    # Print summary report
    print("\n" + "=" * 60)
    print("           SUMMARY REPORT")
    print("=" * 60)
    print(f"  Total rows processed:                  {total_rows}")
    print(f"  Rows with SKU:                         {rows_with_sku}")
    print(f"  Successful matches from content-master: {content_matches}")
    print(f"  Successful matches from master:         {master_matches}")
    print(f"  Rows with no matches at all:            {no_match_rows}")
    print("=" * 60)


if __name__ == "__main__":
    import os

    base_dir = os.path.dirname(os.path.abspath(__file__))

    products_file = os.path.join(base_dir, "products.xlsx")
    content_master_file = os.path.join(base_dir, "content_master.xlsx")
    master_file = os.path.join(base_dir, "gs.xlsx")
    output_file = os.path.join(base_dir, "new_product.xlsx")

    populate_products(products_file, content_master_file, master_file, output_file)
